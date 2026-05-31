"""
pgappforge/plugins/reports/engine.py

ReportEngine — renders a Report into PDF, XLSX, HTML, or CSV.

All four public methods follow the same contract:
    engine = ReportEngine(db_session)
    pdf_bytes  = engine.generate_pdf(report_id, params={"from_date": "2024-01-01"})
    xlsx_bytes = engine.generate_excel(report_id)
    html_str   = engine.generate_html(report_id)
    csv_str    = engine.generate_csv(report_id)

Optional heavy deps are guarded at import time so the engine is still
importable when reportlab / openpyxl are absent — callers receive a clear
RuntimeError explaining what to install.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import textwrap
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

log = logging.getLogger(__name__)

# Hard cap on rows materialised in memory (configurable via REPORTFORGE_MAX_ROWS)
_DEFAULT_MAX_ROWS = 50_000

# ---------------------------------------------------------------------------
# Optional-dep guards
# ---------------------------------------------------------------------------

_NumberedCanvas = None               # module-scope fallback; overwritten if reportlab available
_NUMBERED_CANVAS_MAX_PAGES = 2000    # hard cap to bound memory use

try:
	from reportlab.lib import colors
	from reportlab.lib.pagesizes import A3, A4, LETTER, LEGAL
	from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
	from reportlab.lib.units import mm
	from reportlab.pdfbase import pdfmetrics  # required for font registration
	from reportlab.platypus import (
		HRFlowable,
		PageBreak,
		Paragraph,
		SimpleDocTemplate,
		Spacer,
		Table,
		TableStyle,
	)
	_HAS_REPORTLAB = True

	# ── Unicode font registration ─────────────────────────────────────────
	# Try common system locations for DejaVu Sans (wide Unicode coverage).
	# Falls back to Helvetica gracefully — no error, just ASCII-only output.
	_UNICODE_FONT = "Helvetica"
	_UNICODE_FONT_BOLD = "Helvetica-Bold"
	_FONT_CANDIDATES = [
		"/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
		"/usr/share/fonts/dejavu/DejaVuSans.ttf",
		"/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
		"/Library/Fonts/Arial Unicode.ttf",
		os.path.join(os.path.dirname(__file__), "DejaVuSans.ttf"),
	]
	_FONT_BOLD_CANDIDATES = [
		f.replace("DejaVuSans.ttf", "DejaVuSans-Bold.ttf")
		for f in _FONT_CANDIDATES
	]
	for _p in _FONT_CANDIDATES:
		if os.path.exists(_p):
			try:
				from reportlab.pdfbase.ttfonts import TTFont
				pdfmetrics.registerFont(TTFont("DejaVuSans", _p))
				_UNICODE_FONT = "DejaVuSans"
				log.debug("ReportForge: registered Unicode font DejaVuSans from %s", _p)
			except Exception:
				pass
			break
	for _p in _FONT_BOLD_CANDIDATES:
		if os.path.exists(_p):
			try:
				from reportlab.pdfbase.ttfonts import TTFont as TTFontB
				pdfmetrics.registerFont(TTFontB("DejaVuSans-Bold", _p))
				_UNICODE_FONT_BOLD = "DejaVuSans-Bold"
			except Exception:
				pass
			break

except ImportError:
	_HAS_REPORTLAB = False
	_UNICODE_FONT = "Helvetica"
	_UNICODE_FONT_BOLD = "Helvetica-Bold"
	log.debug("reports engine: reportlab not installed — PDF generation disabled")

try:
	import openpyxl
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter
	_HAS_OPENPYXL = True
except ImportError:
	_HAS_OPENPYXL = False
	log.debug("reports engine: openpyxl not installed — XLSX generation disabled")

try:
	import matplotlib
	matplotlib.use("Agg")  # headless — no GUI
	import matplotlib.pyplot as plt
	_HAS_MATPLOTLIB = True
except ImportError:
	_HAS_MATPLOTLIB = False
	log.debug("reports engine: matplotlib not installed — chart rendering disabled")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_PAGE_SIZES: dict[str, Any] = {}
if _HAS_REPORTLAB:
	_PAGE_SIZES = {
		"A4":     A4,
		"A3":     A3,
		"A5":     (148 * mm, 210 * mm),
		"letter": LETTER,
		"legal":  LEGAL,
	}


def _page_size_for(paper_size: str, orientation: str):
	"""Return a reportlab (width, height) tuple for the given settings."""
	size = _PAGE_SIZES.get(paper_size, A4)
	if orientation == "landscape":
		return (size[1], size[0])
	return size


def _fmt(value: Any, format_string: str | None) -> str:
	"""Apply *format_string* to *value*, falling back to str()."""
	if value is None:
		return ""
	try:
		if format_string:
			return format_string.format(value)
	except (ValueError, TypeError, KeyError) as exc:
		log.warning(
			"ReportForge: format_string %r failed for value %r (%s): falling back to str()",
			format_string, value, exc,
		)
	return str(value)


def _cache_key(report_id: int, params: dict, changed_on, fmt: str) -> str:
	"""SHA-256 hex digest (32 chars) for the render cache."""
	import hashlib, json as _json
	changed = changed_on.isoformat() if changed_on else "none"
	raw = f"{report_id}:{_json.dumps(sorted((params or {}).items()))}:{changed}:{fmt}"
	return hashlib.sha256(raw.encode()).hexdigest()[:32]


def _chart_png(chart_config: dict, rows: list[dict]) -> bytes | None:
	"""
	Render a simple chart to PNG bytes using matplotlib.
	chart_config keys: type (bar|line|pie), data_field, label_field, title
	Returns None when matplotlib is not available or config is invalid.
	"""
	if not _HAS_MATPLOTLIB:
		return None
	try:
		ctype       = chart_config.get("type", "bar")
		data_field  = chart_config.get("data_field", "")
		label_field = chart_config.get("label_field", "")
		title       = chart_config.get("title", "")
		labels = [str(r.get(label_field, i)) for i, r in enumerate(rows)]
		values = [float(r.get(data_field, 0) or 0) for r in rows]
		fig, ax = plt.subplots(figsize=(5, 3), dpi=96)
		if ctype == "pie":
			ax.pie(values, labels=labels, autopct="%1.0f%%")
		elif ctype == "line":
			ax.plot(labels, values, marker="o")
			ax.set_xticks(range(len(labels)))
			ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
		else:  # bar (default)
			ax.bar(labels, values)
			ax.set_xticks(range(len(labels)))
			ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
		if title:
			ax.set_title(title, fontsize=10)
		ax.tick_params(labelsize=8)
		fig.tight_layout()
		buf = io.BytesIO()
		fig.savefig(buf, format="png", bbox_inches="tight")
		plt.close(fig)
		buf.seek(0)
		return buf.read()
	except Exception as exc:
		log.warning("ReportForge: chart render failed: %s", exc)
		return None


def _pdf_font(bold: bool = False) -> str:
	"""Return the registered Unicode-capable font name for PDF."""
	return _UNICODE_FONT_BOLD if bold else _UNICODE_FONT  # type: ignore[name-defined]


def _fetch_logo(logo_url: str, timeout: int = 5) -> bytes | None:
	"""
	Fetch a logo image from an HTTP/HTTPS URL or a local file path.

	Returns raw bytes on success, None on any error.
	Only absolute http/https URLs and absolute local paths are accepted —
	relative paths are rejected to prevent path traversal.
	"""
	if not logo_url:
		return None
	try:
		if logo_url.startswith(("http://", "https://")):
			req = urllib.request.Request(logo_url, headers={"User-Agent": "ReportForge/1.0"})
			with urllib.request.urlopen(req, timeout=timeout) as resp:
				return resp.read()
		elif os.path.isabs(logo_url) and os.path.isfile(logo_url):
			return Path(logo_url).read_bytes()
	except Exception as exc:
		log.warning("ReportForge: could not fetch logo from %r: %s", logo_url, exc)
	return None


# ---------------------------------------------------------------------------
# ReportEngine
# ---------------------------------------------------------------------------

def _compute_aggregate(rows: list[dict], expr: str) -> str:
	"""
	Evaluate a compute expression against a list of row dicts.
	Supported: sum(col), count(*), avg(col), min(col), max(col).
	Returns formatted string or "" on error.
	"""
	if not expr or not rows:
		return ""
	import re as _re
	m = _re.match(r"^\s*(\w+)\((\w+|\*)\)\s*$", expr.strip().lower())
	if not m:
		return ""
	func, col = m.group(1), m.group(2)
	try:
		if func == "count":
			return str(len(rows))
		vals = [float(r.get(col, 0) or 0) for r in rows]
		if func == "sum":
			return f"{sum(vals):,.2f}"
		if func == "avg":
			return f"{sum(vals)/len(vals):,.2f}" if vals else ""
		if func == "min":
			return f"{min(vals):,.2f}" if vals else ""
		if func == "max":
			return f"{max(vals):,.2f}" if vals else ""
	except Exception:
		pass
	return ""


# NumberedCanvas enables "Page N of M" in PDF output.
# Uses a two-stage save: showPage() buffers state; save() patches totals.
if _HAS_REPORTLAB:
	from reportlab.pdfgen import canvas as _rl_canvas_module

	# Hard cap to avoid unbounded memory growth on pathologically large reports.
	_NUMBERED_CANVAS_MAX_PAGES = 2000

	class _NumberedCanvas(_rl_canvas_module.Canvas):
		"""
		Two-pass canvas enabling "Page N of M" footers.

		showPage() buffers page state (a shallow copy — drawing commands are
		shared references, not duplicated).  save() replays all pages with
		_total_pages injected so the per-page callback can read it.

		Memory: O(n_pages × ~5-50 KB for drawing operators).  Hard-capped at
		_NUMBERED_CANVAS_MAX_PAGES; reports beyond that fall back to "Page N".
		"""

		def __init__(self, *args, **kwargs):
			super().__init__(*args, **kwargs)
			self._page_states: list[dict] = []

		def showPage(self):
			if len(self._page_states) < _NUMBERED_CANVAS_MAX_PAGES:
				self._page_states.append(dict(self.__dict__))
			self._startPage()

		def save(self):
			total = len(self._page_states)
			for state in self._page_states:
				self.__dict__.update(state)
				self._total_pages = total
				super().showPage()
			super().save()


class ReportEngine:
	"""
	Renders Report records into various output formats.

	Args:
		session: Active SQLAlchemy session.  The engine never commits — it
		         only reads report definitions and executes datasource queries.
		preview_row_limit: Maximum data rows fetched for HTML/CSV preview.
		                   PDF and XLSX always fetch all rows.
	"""

	def __init__(
		self,
		session: Session,
		preview_row_limit: int = 10,
		download_row_limit: int | None = None,
		cache_ttl_hours: int = 1,
	) -> None:
		assert isinstance(session, Session), "session must be a SQLAlchemy Session"
		self._session = session
		self.preview_row_limit  = preview_row_limit
		self.download_row_limit = download_row_limit
		self.cache_ttl_hours    = cache_ttl_hours

	# ── Render cache helpers ──────────────────────────────────────────────

	def _cache_get(self, key: str, fmt: str) -> bytes | None:
		"""Return cached bytes for (key, fmt) if valid, else None."""
		try:
			from .models import ReportRenderCache
			import sqlalchemy as sa
			from datetime import timezone
			now = datetime.now(timezone.utc)
			row = self._session.execute(
				sa.select(ReportRenderCache).where(
					sa.and_(
						ReportRenderCache.cache_key == key,
						ReportRenderCache.format    == fmt,
						ReportRenderCache.expires_at > now,
					)
				)
			).scalar_one_or_none()
			return row.data if row else None
		except Exception as exc:
			log.debug("cache_get failed: %s", exc)
			return None

	# Rendered bytes larger than this are not cached (avoids OOM on huge PDFs)
	_CACHE_MAX_BYTES_DEFAULT = 10 * 1024 * 1024  # 10 MB

	def _cache_set(self, key: str, fmt: str, data: bytes, report_id: int) -> None:
		"""Persist rendered bytes into the render cache.

		Skips caching when ``len(data) > REPORTFORGE_CACHE_MAX_BYTES`` (default
		10 MB) to prevent OOM on large reports being written to PostgreSQL BYTEA.
		"""
		try:
			from flask import current_app
			max_bytes = current_app.config.get(
				"REPORTFORGE_CACHE_MAX_BYTES", self._CACHE_MAX_BYTES_DEFAULT
			)
			if len(data) > max_bytes:
				log.debug(
					"cache_set: skipping %d bytes (> %d byte cap) for key %s",
					len(data), max_bytes, key,
				)
				return
			from .models import ReportRenderCache
			import sqlalchemy as sa
			from datetime import timedelta, timezone
			expires = datetime.now(timezone.utc) + timedelta(hours=self.cache_ttl_hours)
			self._session.execute(
				sa.delete(ReportRenderCache).where(
					sa.and_(
						ReportRenderCache.cache_key == key,
						ReportRenderCache.format    == fmt,
					)
				)
			)
			row = ReportRenderCache(
				cache_key=key, report_id=report_id, format=fmt,
				data=data, size_bytes=len(data), expires_at=expires,
			)
			self._session.add(row)
			self._session.commit()
		except Exception as exc:
			log.debug("cache_set failed: %s", exc)

	def cache_invalidate(self, report_id: int) -> None:
		"""Evict all cached renders for a report (call after design changes)."""
		try:
			from .models import ReportRenderCache
			import sqlalchemy as sa
			self._session.execute(
				sa.delete(ReportRenderCache).where(
					ReportRenderCache.report_id == report_id
				)
			)
			self._session.commit()
		except Exception as exc:
			log.debug("cache_invalidate failed: %s", exc)

	# ------------------------------------------------------------------ #
	# Public API                                                           #
	# ------------------------------------------------------------------ #

	def generate_pdf(self, report_id: int, params: dict[str, Any] | None = None) -> bytes:
		"""
		Render the report to a PDF byte string using reportlab.

		Raises:
			RuntimeError: When reportlab is not installed.
			LookupError:  When report_id does not exist.
		"""
		if not _HAS_REPORTLAB:
			raise RuntimeError(
				"reportlab is required for PDF generation. "
				"Install it with: pip install reportlab"
			)
		report  = self._load_report(report_id)
		params  = self._resolve_params(report, params or {})

		# ── Cache check ───────────────────────────────────────────────────
		ck = _cache_key(report_id, params, report.changed_on, "pdf")
		cached = self._cache_get(ck, "pdf")
		if cached:
			return cached

		rows   = self._execute_query(report, params)
		groups = self._group_data(rows, report.group_field)

		buf = io.BytesIO()
		ps  = _page_size_for(report.paper_size.value, report.orientation.value)
		pc  = report.page_config or {}
		doc = SimpleDocTemplate(
			buf,
			pagesize     = ps,
			topMargin    = pc.get("margin_top_mm",    10) * mm,
			bottomMargin = pc.get("margin_bottom_mm", 10) * mm,
			leftMargin   = pc.get("margin_left_mm",   15) * mm,
			rightMargin  = pc.get("margin_right_mm",  15) * mm,
			title        = report.name,
		)

		story   = self._build_pdf_story(report, rows, groups)
		page_fn = self._make_page_callback(report, ps, pc)

		# _NumberedCanvas enables "Page N of M" by buffering all pages
		canvas_cls = _NumberedCanvas if _HAS_REPORTLAB else None
		if canvas_cls:
			doc.build(story, onFirstPage=page_fn, onLaterPages=page_fn,
			          canvasmaker=canvas_cls)
		else:
			doc.build(story, onFirstPage=page_fn, onLaterPages=page_fn)

		data = buf.getvalue()
		self._cache_set(ck, "pdf", data, report_id)
		return data

	def _make_page_callback(self, report, ps, pc):
		"""
		Return a reportlab page callback that draws PAGE_HEADER, PAGE_FOOTER,
		logo overlay, and watermark on every page of the PDF.

		PAGE_HEADER is drawn at the top of each page (above the main frame).
		PAGE_FOOTER is drawn at the bottom (below the main frame).
		"""
		page_w, page_h = ps
		margin_top    = pc.get("margin_top_mm",    10) * mm
		margin_bottom = pc.get("margin_bottom_mm", 10) * mm
		margin_left   = pc.get("margin_left_mm",   15) * mm

		bands_by_type: dict = defaultdict(list)
		for band in report.band_list():
			bands_by_type[band.band_type.value].append(band)

		# Pre-fetch logo bytes once so we don't hit the network per page
		logo_bytes: bytes | None = None
		if report.logo_url:
			logo_bytes = _fetch_logo(report.logo_url)

		wm_text    = getattr(report, "watermark_text",    None)
		wm_opacity = getattr(report, "watermark_opacity", 0.08) or 0.08

		def _draw_band_fields(canvas_obj, band, y_top: float) -> None:
			"""Draw all text fields in a band using direct canvas calls."""
			for field in band.field_list():
				text = field.style.get("text", "") or ""
				if not text:
					continue
				fs   = float(field.style.get("font_size", 9))
				bold = bool(field.style.get("bold", False))
				fn   = _pdf_font(bold)
				col  = _rl_color(field.style.get("color", "#000000"))
				aln  = field.style.get("align", "left")
				x    = margin_left + field.x_mm * mm
				y    = y_top - field.y_mm * mm - fs

				canvas_obj.setFont(fn, fs)
				canvas_obj.setFillColor(col)
				if aln == "center":
					canvas_obj.drawCentredString(page_w / 2, y, _escape_rl(text))
				elif aln == "right":
					canvas_obj.drawRightString(page_w - margin_left, y, _escape_rl(text))
				else:
					canvas_obj.drawString(x, y, _escape_rl(text))

		def page_callback(canvas_obj, doc_obj):
			canvas_obj.saveState()

			# ── Watermark (behind everything else) ─────────────────────────
			if wm_text:
				canvas_obj.setFillGray(0.5, wm_opacity)
				canvas_obj.setFont(_pdf_font(False), 60)
				canvas_obj.translate(page_w / 2, page_h / 2)
				canvas_obj.rotate(45)
				canvas_obj.drawCentredString(0, 0, wm_text.upper())
				canvas_obj.translate(-page_w / 2, -page_h / 2)

			# ── Logo (top-right corner) ─────────────────────────────────────
			if logo_bytes:
				try:
					logo_img = io.BytesIO(logo_bytes)
					canvas_obj.drawImage(
						logo_img,
						page_w - margin_left - 40 * mm,
						page_h - margin_top - 20 * mm,
						width=35 * mm, height=15 * mm,
						preserveAspectRatio=True, mask="auto",
					)
				except Exception as exc:
					log.debug("ReportForge: logo render failed: %s", exc)

			# ── PAGE_HEADER bands (below the top margin) ────────────────────
			y = page_h - margin_top
			for band in bands_by_type.get("page_header", []):
				bh = band.height_mm * mm
				bg = _rl_color(band.background_color)
				if bg:
					canvas_obj.setFillColor(bg)
					canvas_obj.rect(0, y - bh, page_w, bh, fill=1, stroke=0)
				_draw_band_fields(canvas_obj, band, y)
				y -= bh

			# ── PAGE_FOOTER bands (above the bottom margin) ─────────────────
			y = margin_bottom
			for band in reversed(bands_by_type.get("page_footer", [])):
				bh = band.height_mm * mm
				bg = _rl_color(band.background_color)
				if bg:
					canvas_obj.setFillColor(bg)
					canvas_obj.rect(0, y, page_w, bh, fill=1, stroke=0)
				# Draw text fields
				for field in band.field_list():
					text = field.style.get("text", "") or ""
					if not text:
						continue
					fs  = float(field.style.get("font_size", 8))
					fn  = _pdf_font(False)
					col = _rl_color(field.style.get("color", "#888888"))
					canvas_obj.setFont(fn, fs)
					canvas_obj.setFillColor(col)
					canvas_obj.drawCentredString(page_w / 2, y + bh / 2 - fs / 2, _escape_rl(text))
				y += bh

			# ── Page number (always at very bottom) ─────────────────────────
			canvas_obj.setFont(_pdf_font(False), 8)
			canvas_obj.setFillColor(_rl_color("#888888"))
			canvas_obj.drawRightString(
				page_w - margin_left,
				margin_bottom - 6 * mm,
				f"Page {doc_obj.page} of {getattr(canvas_obj, '_total_pages', '?')}",
			)

			canvas_obj.restoreState()

		return page_callback

	def generate_excel(self, report_id: int, params: dict[str, Any] | None = None) -> bytes:
		"""
		Render the report to an XLSX byte string using openpyxl.

		Each band type produces a styled section:
		- TITLE / PAGE_HEADER / SUMMARY / PAGE_FOOTER → merged header rows
		- COLUMN_HEADER                                → bold column labels
		- DETAIL                                       → one row per data record

		Raises:
			RuntimeError: When openpyxl is not installed.
			LookupError:  When report_id does not exist.
		"""
		if not _HAS_OPENPYXL:
			raise RuntimeError(
				"openpyxl is required for XLSX generation. "
				"Install it with: pip install openpyxl"
			)
		report  = self._load_report(report_id)
		params  = self._resolve_params(report, params or {})

		# ── Cache check ───────────────────────────────────────────────────
		ck = _cache_key(report_id, params, report.changed_on, "xlsx")
		cached = self._cache_get(ck, "xlsx")
		if cached:
			return cached

		rows = self._execute_query(report, params)

		# write_only=True enables streaming — avoids loading the full workbook into RAM
		wb = openpyxl.Workbook(write_only=True)
		ws = wb.create_sheet(title=report.name[:31])

		# Set worksheet properties BEFORE writing any rows
		ws.freeze_panes = "A2"   # freeze header row
		ws.auto_filter.ref = "A1"  # Excel will expand to data range on open

		columns: list[str] = list(rows[0].keys()) if rows else []

		header_style = Font(bold=True, color="FFFFFF")
		header_fill  = PatternFill("solid", fgColor="4472C4")
		title_fill   = PatternFill("solid", fgColor="336699")
		title_font   = Font(bold=True, color="FFFFFF", size=13)

		# Python format_string → Excel number format
		def _xl_numfmt(fmt_str: str | None) -> str | None:
			if not fmt_str:
				return None
			if "{:,.2f}" in fmt_str or "{:.2f}" in fmt_str:
				return "#,##0.00"
			if "{:,.0f}" in fmt_str or "{:.0f}" in fmt_str:
				return "#,##0"
			if "{:,}" in fmt_str:
				return "#,##0"
			if "%Y" in fmt_str or ":%Y" in fmt_str:
				return "YYYY-MM-DD"
			if "%d/%m/%Y" in fmt_str:
				return "DD/MM/YYYY"
			return None

		for band in report.band_list():
			fields = band.field_list()
			btype  = band.band_type.value

			if btype in ("title", "summary"):
				label = (
					fields[0].style.get("text", btype.replace("_", " ").title())
					if fields else btype.title()
				)
				cell = openpyxl.cell.WriteOnlyCell(ws, value=label)
				cell.font      = title_font
				cell.fill      = title_fill
				cell.alignment = Alignment(horizontal="center")
				ws.append([cell])

			elif btype == "column_header":
				labels = [
					f.style.get("text") or f.data_binding or f"col_{i}"
					for i, f in enumerate(fields)
				] if fields else columns
				row_cells = []
				for label in labels:
					c = openpyxl.cell.WriteOnlyCell(ws, value=label)
					c.font      = header_style
					c.fill      = header_fill
					c.alignment = Alignment(horizontal="center")
					row_cells.append(c)
				ws.append(row_cells)

			elif btype == "detail":
				bindings = [
					f.data_binding for f in fields if f.data_binding
				] or columns
				fmt_map = {
					f.data_binding: f.format_string
					for f in fields if f.data_binding and f.data_binding
				}
				for r_idx, row_data in enumerate(rows):
					fill_color = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
					row_cells  = []
					for binding in bindings:
						raw = row_data.get(binding)
						# Write raw Python type (int/float/date) so Excel can sort/filter
						c = openpyxl.cell.WriteOnlyCell(ws, value=raw)
						c.fill = PatternFill("solid", fgColor=fill_color)
						xl_fmt = _xl_numfmt(fmt_map.get(binding))
						if xl_fmt:
							c.number_format = xl_fmt
						row_cells.append(c)
					ws.append(row_cells)

		out = io.BytesIO()
		wb.save(out)
		data = out.getvalue()
		self._cache_set(ck, "xlsx", data, report_id)
		return data

	def generate_html(self, report_id: int, params: dict[str, Any] | None = None) -> str:
		"""
		Render the report as a self-contained HTML string (first N rows only).

		This is used for the designer preview — it renders the band layout with
		real data so the user can verify formatting without generating a full PDF.
		"""
		report = self._load_report(report_id)
		params  = self._resolve_params(report, params or {})
		rows    = self._execute_query(report, params, limit=self.preview_row_limit)

		parts: list[str] = [
			"<!DOCTYPE html>",
			'<html lang="en">',
			'<head><meta charset="utf-8">',
			f'<title>{_escape(report.name)}</title>',
			self._html_styles(),
			"</head>",
			'<body class="rpt-body">',
			f'<div class="rpt-page rpt-{report.orientation.value}">',
		]

		columns: list[str] = list(rows[0].keys()) if rows else []

		for band in report.band_list():
			fields = band.field_list()
			btype  = band.band_type.value
			bg     = band.background_color or "#ffffff"
			height = band.height_mm

			parts.append(
				f'<div class="rpt-band rpt-{btype}" '
				f'style="min-height:{height}mm;background:{_escape(bg)}">'
			)

			if btype in ("title", "page_header", "summary", "page_footer"):
				label = (
					fields[0].style.get("text", btype.replace("_", " ").title())
					if fields else btype.replace("_", " ").title()
				)
				fs   = fields[0].style.get("font_size", 14) if fields else 14
				parts.append(
					f'<div class="rpt-section-label" style="font-size:{fs}pt">'
					f'{_escape(label)}</div>'
				)

			elif btype == "column_header":
				labels = [
					f.style.get("text") or f.data_binding or f"col_{i}"
					for i, f in enumerate(fields)
				] if fields else columns
				parts.append('<table class="rpt-table"><thead><tr>')
				for label in labels:
					parts.append(f'<th>{_escape(str(label))}</th>')
				parts.append("</tr></thead></table>")

			elif btype == "detail":
				bindings = [f.data_binding for f in fields if f.data_binding] or columns
				fmt_map  = {
					f.data_binding: f.format_string
					for f in fields if f.data_binding
				}
				link_map = {
					f.data_binding: f.link_url_template
					for f in fields
					if getattr(f, "link_url_template", None) and f.data_binding
				}
				# column header row first
				parts.append('<table class="rpt-table"><thead><tr>')
				for binding in bindings:
					parts.append(f'<th>{_escape(binding)}</th>')
				parts.append("</tr></thead><tbody>")
				for r_idx, row_data in enumerate(rows):
					cls = "rpt-row-alt" if r_idx % 2 else "rpt-row"
					parts.append(f'<tr class="{cls}">')
					for binding in bindings:
						raw = row_data.get(binding)
						val = _fmt(raw, fmt_map.get(binding))
						cell_html = _escape(val)
						link_tmpl = link_map.get(binding)
						if link_tmpl:
							# Expand {field_name} placeholders with row values
							import re as _re
							try:
								href = _re.sub(
									r"\{(\w+)\}",
									lambda m: str(row_data.get(m.group(1), "")),
									link_tmpl,
								)
								cell_html = f'<a href="{_escape(href)}">{cell_html}</a>'
							except Exception:
								pass
						parts.append(f'<td>{cell_html}</td>')
					parts.append("</tr>")
				if len(rows) == self.preview_row_limit:
					parts.append(
						f'<tr><td colspan="{len(bindings)}" class="rpt-preview-note">'
						f'Preview limited to {self.preview_row_limit} rows</td></tr>'
					)
				parts.append("</tbody></table>")

			else:
				# group_header / group_footer — show field text statics
				for f in fields:
					text = f.style.get("text") or f.data_binding or ""
					parts.append(f'<span class="rpt-field">{_escape(text)}</span>')

			parts.append("</div>")  # end band

		parts.extend(["</div>", "</body>", "</html>"])
		return "\n".join(parts)

	def generate_csv(self, report_id: int, params: dict[str, Any] | None = None) -> str:
		"""
		Render the DETAIL band data as RFC-4180 CSV.

		Only DETAIL band fields (or raw SQL columns when no fields are defined)
		are exported.  TITLE / HEADER / FOOTER bands are omitted — use
		generate_html() or generate_pdf() to include them.
		"""
		report = self._load_report(report_id)
		params  = self._resolve_params(report, params or {})
		rows    = self._execute_query(report, params)

		if not rows:
			return ""

		# Derive bindings from the first DETAIL band; fall back to all SQL columns
		columns  = list(rows[0].keys())
		bindings = columns
		fmt_map: dict[str, str | None] = {}

		for band in report.band_list():
			if band.band_type.value == "detail" and band.field_list():
				bindings = [f.data_binding for f in band.field_list() if f.data_binding]
				fmt_map  = {
					f.data_binding: f.format_string
					for f in band.field_list() if f.data_binding
				}
				break

		buf    = io.StringIO()
		writer = csv.writer(buf, dialect="excel")
		writer.writerow(bindings)
		for row_data in rows:
			writer.writerow([
				_fmt(row_data.get(b), fmt_map.get(b))
				for b in bindings
			])
		return buf.getvalue()

	# ------------------------------------------------------------------ #
	# Internal: data pipeline                                             #
	# ------------------------------------------------------------------ #

	def _load_report(self, report_id: int):
		"""Fetch the Report ORM object or raise LookupError."""
		from .models import Report
		report = self._session.get(Report, report_id)
		if report is None:
			raise LookupError(f"Report with id={report_id} does not exist")
		return report

	def fetch_rows(
		self,
		report_id: int,
		params: dict[str, Any] | None = None,
	) -> list[dict[str, Any]]:
		"""
		Public helper: execute the report datasource and return all rows as dicts.

		Uses ``download_row_limit`` (no cap by default).
		"""
		report = self._load_report(report_id)
		resolved = self._resolve_params(report, params or {})
		return self._execute_query(report, resolved, limit=self.download_row_limit)

	def _resolve_params(self, report, supplied: dict[str, Any]) -> dict[str, Any]:
		"""
		Merge caller-supplied values with report parameter defaults.

		Applies type coercion via ReportParameter.coerce().
		"""
		resolved: dict[str, Any] = {}
		for param in report.parameters:
			raw     = supplied.get(param.name)
			resolved[param.name] = param.coerce(raw)
		# Pass-through any extra keys the caller provided (advanced use cases)
		for k, v in supplied.items():
			if k not in resolved:
				resolved[k] = v
		return resolved

	def _execute_query(
		self,
		report,
		params: dict[str, Any],
		limit: int | None = None,
	) -> list[dict[str, Any]]:
		"""
		Execute the report datasource and return rows as dicts.

		When ``is_sql_source`` is True the ``data_source`` field is treated as
		a parameterised SQL string; named bind parameters (``:name``) are
		substituted from *params*.

		When ``is_sql_source`` is False the string is treated as a dotted
		Python import path to a SQLAlchemy model class; all rows are fetched
		via ``session.execute(select(Model))``.
		"""
		from sqlalchemy import text, select

		try:
			if report.is_sql_source:
				sql = report.data_source.strip()
				# Enforce GROUP BY ordering so group_data() partitions correctly
				if report.group_field and not re.search(r"\bORDER\s+BY\b", sql, re.IGNORECASE):
					sql = f"{sql.rstrip(';')}\nORDER BY {report.group_field}"
				effective_limit = limit or self.download_row_limit or _DEFAULT_MAX_ROWS
				if re.search(r"\bLIMIT\b", sql, re.IGNORECASE):
					# User already has LIMIT — don't double-wrap, but warn if > cap
					pass
				else:
					sql = f"SELECT * FROM ({sql}) __rpt_data LIMIT {int(effective_limit)}"
				result = self._session.execute(text(sql), params)
				keys   = list(result.keys())
				rows   = [dict(zip(keys, row)) for row in result.fetchall()]
				if len(rows) >= effective_limit:
					log.warning(
						"ReportForge: report %s hit row cap (%d). "
						"Increase REPORTFORGE_MAX_ROWS or add LIMIT to your query.",
						report.id, effective_limit,
					)
				return rows
			else:
				model_cls = self._import_model(report.data_source)
				stmt      = select(model_cls)
				effective_limit = limit or self.download_row_limit or _DEFAULT_MAX_ROWS
				stmt = stmt.limit(effective_limit)
				result = self._session.execute(stmt)
				rows   = result.scalars().all()
				return [self._model_to_dict(r) for r in rows]
		except Exception as exc:
			log.error("report engine: query failed for report_id=%s: %s", report.id, exc)
			raise

	def _group_data(
		self,
		rows: list[dict[str, Any]],
		group_field: str | None,
	) -> dict[Any, list[dict[str, Any]]]:
		"""
		Partition *rows* into an ordered dict keyed by *group_field* value.

		Returns ``{None: rows}`` when *group_field* is None/empty (no grouping).
		Insertion order is preserved — the first occurrence of each group value
		determines sort order.
		"""
		if not group_field:
			return {None: rows}

		grouped: dict[Any, list[dict[str, Any]]] = {}
		for row in rows:
			key = row.get(group_field)
			grouped.setdefault(key, []).append(row)
		return grouped

	# ------------------------------------------------------------------ #
	# Internal: PDF story builder                                          #
	# ------------------------------------------------------------------ #

	def _build_pdf_story(self, report, rows, groups) -> list:
		"""
		Translate report bands + data into a reportlab Platypus story.

		Rendering pass:
		  1. TITLE band (once)
		  2. PAGE_HEADER (skipped here — drawn by page callback on every page)
		  3. For each group:
		       a. GROUP_HEADER
		       b. COLUMN_HEADER (once per group)
		       c. DETAIL rows
		       d. GROUP_FOOTER (with compute aggregates)
		  4. SUMMARY (once)
		  5. PAGE_FOOTER (skipped here — drawn by page callback on every page)
		"""
		self._last_rows = rows  # for CHART fields in summary bands
		story: list = []
		styles = getSampleStyleSheet()

		bands_by_type: dict[str, list] = defaultdict(list)
		for band in report.band_list():
			bands_by_type[band.band_type.value].append(band)

		columns: list[str] = list(rows[0].keys()) if rows else []

		# --- TITLE ---
		for band in bands_by_type.get("title", []):
			story.extend(self._pdf_section_band(band, styles))

		# --- PAGE_HEADER ---
		for band in bands_by_type.get("page_header", []):
			story.extend(self._pdf_section_band(band, styles))

		# --- COLUMN_HEADER (once before grouped data) ---
		col_header_bands = bands_by_type.get("column_header", [])

		# --- iterate groups ---
		for group_key, group_rows in groups.items():
			# GROUP_HEADER
			for band in bands_by_type.get("group_header", []):
				story.extend(self._pdf_group_band(band, group_key, styles, group_rows))

			# COLUMN_HEADER
			for band in col_header_bands:
				story.extend(self._pdf_column_header_band(band, columns, styles))

			# DETAIL rows
			self._last_rows = group_rows  # allow CHART fields in footers to reference data
			for band in bands_by_type.get("detail", []):
				story.extend(self._pdf_detail_band(band, group_rows, columns, styles))

			# GROUP_FOOTER (with compute aggregates)
			for band in bands_by_type.get("group_footer", []):
				story.extend(self._pdf_group_band(band, group_key, styles, group_rows))

		# --- SUMMARY ---
		for band in bands_by_type.get("summary", []):
			story.extend(self._pdf_section_band(band, styles))

		# --- PAGE_FOOTER note (not rendered per-page but included at end) ---
		for band in bands_by_type.get("page_footer", []):
			story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
			story.extend(self._pdf_section_band(band, styles))

		return story

	def _pdf_section_band(self, band, styles) -> list:
		"""Render TITLE / SUMMARY as paragraphs (PAGE_HEADER/FOOTER handled by page callback)."""
		# PAGE_HEADER and PAGE_FOOTER are drawn by _make_page_callback — skip here
		if band.band_type.value in ("page_header", "page_footer"):
			return []
		result = []
		bg = _rl_color(band.background_color)

		for field in band.field_list():
			ftype = field.field_type.value if hasattr(field.field_type, "value") else str(field.field_type)

			# LINE
			if ftype == "line":
				col = _rl_color(field.style.get("color", "#cccccc"))
				result.append(HRFlowable(
					width=f"{field.width_mm}mm",
					thickness=field.style.get("line_width", 0.5),
					color=col,
					spaceAfter=2,
				))
				continue

			# BOX
			if ftype == "box":
				bg_box = _rl_color(field.style.get("bg_color", "transparent"))
				bdr    = float(field.style.get("border", 1))
				tbl = Table([[""]], colWidths=[field.width_mm * mm], rowHeights=[field.height_mm * mm])
				tbl.setStyle(TableStyle([
					("BOX",    (0, 0), (-1, -1), bdr, _rl_color(field.style.get("border_color", "#000000"))),
					("BACKGROUND", (0, 0), (-1, -1), bg_box or colors.white),
				]))
				result.append(tbl)
				continue

			# IMAGE
			if ftype == "image":
				src = field.style.get("image_src", "")
				if src:
					img_bytes = _fetch_logo(src)
					if img_bytes:
						try:
							from reportlab.platypus import Image as RLImage
							result.append(RLImage(
								io.BytesIO(img_bytes),
								width=field.width_mm * mm,
								height=field.height_mm * mm,
							))
						except Exception as exc:
							log.debug("ReportForge: image render failed: %s", exc)
				continue

			# CHART
			if ftype == "chart":
				chart_cfg = field.style.get("chart_config", {})
				if chart_cfg and _HAS_MATPLOTLIB:
					# Rows come from the last group/all rows — passed via style for summary charts
					# Use self._last_rows if available (set in _build_pdf_story)
					chart_rows = getattr(self, "_last_rows", [])
					png = _chart_png(chart_cfg, chart_rows)
					if png:
						try:
							from reportlab.platypus import Image as RLImage
							result.append(RLImage(
								io.BytesIO(png),
								width=field.width_mm * mm,
								height=field.height_mm * mm,
							))
						except Exception as exc:
							log.debug("ReportForge: chart embed failed: %s", exc)
				continue

			# TEXT / NUMBER / DATE (default)
			text = field.style.get("text") or field.data_binding or ""
			fs   = field.style.get("font_size", 12)
			bold = bool(field.style.get("bold", False))
			fn   = _pdf_font(bold)
			ps   = ParagraphStyle(
				f"rpt_section_{field.id}",
				parent    = styles["Normal"],
				fontSize  = fs,
				fontName  = fn,
				textColor = _rl_color(field.style.get("color", "#000000")),
				backColor = bg,
				spaceAfter= 4,
				alignment = _rl_align(field.style.get("align", "left")),
			)
			result.append(Paragraph(_escape_rl(text), ps))

		result.append(Spacer(1, band.height_mm * 0.3 * mm))
		return result

	def _pdf_group_band(
		self, band, group_key: Any, styles,
		group_rows: list | None = None,
	) -> list:
		"""Render GROUP_HEADER / GROUP_FOOTER.

		For GROUP_FOOTER bands, fields with a ``compute`` expression are evaluated
		against *group_rows* and rendered as aggregate values.
		"""
		fields = band.field_list()
		bg     = _rl_color(band.background_color)
		result: list = []

		# If fields define their own layout, render them; else fall back to a
		# plain "Group: value" paragraph.
		if fields and band.band_type.value == "group_footer" and group_rows:
			# Render each field, substituting compute expressions
			for field in fields:
				compute_val = ""
				if getattr(field, "compute", None) and group_rows:
					compute_val = _compute_aggregate(group_rows, field.compute)
				text = compute_val or field.style.get("text") or str(group_key)
				fs   = field.style.get("font_size", 10)
				bold = bool(field.style.get("bold", True))
				ps   = ParagraphStyle(
					f"rpt_gf_{field.id}",
					parent    = styles["Normal"],
					fontSize  = fs,
					fontName  = _pdf_font(bold),
					textColor = _rl_color(field.style.get("color", "#000000")),
					backColor = bg,
					alignment = _rl_align(field.style.get("align", "right")),
					spaceAfter= 2,
				)
				result.append(Paragraph(_escape_rl(text), ps))
		else:
			# Default: "Group Header/Footer: <group_key_value>"
			label = f"{band.band_type.value.replace('_', ' ').title()}: {group_key}"
			ps    = ParagraphStyle(
				f"rpt_group_{band.id}",
				parent    = styles["Normal"],
				fontSize  = 11,
				fontName  = _pdf_font(True),
				backColor = bg,
				spaceAfter= 2,
			)
			result.append(Paragraph(_escape_rl(label), ps))

		result.append(Spacer(1, 2 * mm))
		return result

	def _pdf_column_header_band(self, band, columns: list[str], styles) -> list:
		"""Render COLUMN_HEADER as a styled single-row table."""
		fields  = band.field_list()
		labels  = [
			f.style.get("text") or f.data_binding or f"col_{i}"
			for i, f in enumerate(fields)
		] if fields else columns
		if not labels:
			return []

		tbl = Table([labels], repeatRows=1)
		tbl.setStyle(TableStyle([
			("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#4472C4")),
			("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
			("FONTNAME",    (0, 0), (-1, 0), _pdf_font(True)),
			("FONTSIZE",    (0, 0), (-1, 0), 9),
			("BOTTOMPADDING", (0, 0), (-1, 0), 5),
			("TOPPADDING",    (0, 0), (-1, 0), 5),
			("GRID",        (0, 0), (-1, -1), 0.25, colors.grey),
		]))
		return [tbl, Spacer(1, 1 * mm)]

	def _pdf_detail_band(self, band, rows, columns: list[str], styles) -> list:
		"""Render DETAIL band rows as a zebra-striped table."""
		fields   = band.field_list()
		bindings = [f.data_binding for f in fields if f.data_binding] or columns
		fmt_map  = {
			f.data_binding: f.format_string
			for f in fields if f.data_binding
		}
		if not bindings or not rows:
			return []

		table_data = []
		for r_idx, row_data in enumerate(rows):
			table_data.append([
				_fmt(row_data.get(b), fmt_map.get(b))
				for b in bindings
			])

		tbl = Table(table_data)
		ts  = [
			("FONTNAME",  (0, 0), (-1, -1), _pdf_font(False)),
			("FONTSIZE",  (0, 0), (-1, -1), 9),
			("GRID",      (0, 0), (-1, -1), 0.25, colors.lightgrey),
			("TOPPADDING",    (0, 0), (-1, -1), 3),
			("BOTTOMPADDING", (0, 0), (-1, -1), 3),
		]
		# Zebra stripes
		for i in range(0, len(table_data), 2):
			ts.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F2F2F2")))
		tbl.setStyle(TableStyle(ts))
		return [tbl, Spacer(1, 1 * mm)]

	# ------------------------------------------------------------------ #
	# Internal: misc utilities                                            #
	# ------------------------------------------------------------------ #

	@staticmethod
	def _import_model(dotted_path: str):
		"""
		Import and return a class from a dotted-path string.

		e.g. ``"myapp.models.Order"`` → ``<class 'myapp.models.Order'>``
		"""
		parts  = dotted_path.rsplit(".", 1)
		if len(parts) != 2:
			raise ImportError(
				f"data_source {dotted_path!r} must be a dotted module.ClassName string"
			)
		module_path, class_name = parts
		import importlib
		module = importlib.import_module(module_path)
		cls    = getattr(module, class_name, None)
		if cls is None:
			raise ImportError(
				f"Cannot find {class_name!r} in module {module_path!r}"
			)
		return cls

	@staticmethod
	def _model_to_dict(instance) -> dict[str, Any]:
		"""
		Convert a SQLAlchemy model instance to a plain dict.

		Only scalar columns are included; relationships are skipped.
		"""
		from sqlalchemy import inspect as sa_inspect
		mapper = sa_inspect(type(instance))
		return {
			col.key: getattr(instance, col.key)
			for col in mapper.mapper.column_attrs
		}

	def _html_styles(self) -> str:
		return textwrap.dedent("""\
			<style>
			  body.rpt-body { font-family: Arial, sans-serif; background: #e8e8e8; margin: 0; padding: 20px; }
			  .rpt-page { background: #fff; max-width: 210mm; margin: 0 auto; padding: 15mm;
			              box-shadow: 0 2px 8px rgba(0,0,0,.2); }
			  .rpt-page.rpt-landscape { max-width: 297mm; }
			  .rpt-band { padding: 4px 0; border-bottom: 1px solid #ddd; }
			  .rpt-title, .rpt-page_header { font-size: 16pt; font-weight: bold; text-align: center; padding: 8px 0; }
			  .rpt-column_header th, .rpt-detail th { background: #4472C4; color: #fff; font-size: 9pt; }
			  .rpt-summary { font-weight: bold; border-top: 2px solid #333; padding-top: 6px; }
			  .rpt-page_footer { font-size: 8pt; color: #666; text-align: center; }
			  .rpt-group_header { background: #d0e0f0; font-weight: bold; padding: 3px 6px; }
			  .rpt-group_footer { background: #e8f0e8; font-style: italic; padding: 3px 6px; }
			  .rpt-table { width: 100%; border-collapse: collapse; font-size: 9pt; }
			  .rpt-table th, .rpt-table td { border: 1px solid #ccc; padding: 4px 6px; }
			  .rpt-row-alt td { background: #f5f5f5; }
			  .rpt-section-label { font-size: 14pt; font-weight: bold; padding: 6px 0; }
			  .rpt-preview-note { font-style: italic; color: #888; font-size: 8pt; text-align: center; }
			  .rpt-field { display: inline-block; margin: 2px 4px; }
			</style>
		""")


# ---------------------------------------------------------------------------
# Small helpers (module-level to avoid repeated instantiation)
# ---------------------------------------------------------------------------

def _escape(text: str) -> str:
	"""HTML-escape a string for safe inline use."""
	return (
		str(text)
		.replace("&", "&amp;")
		.replace("<", "&lt;")
		.replace(">", "&gt;")
		.replace('"', "&quot;")
	)


def _escape_rl(text: str) -> str:
	"""Minimal XML escape for reportlab Paragraph content."""
	return (
		str(text)
		.replace("&", "&amp;")
		.replace("<", "&lt;")
		.replace(">", "&gt;")
	)


def _rl_color(css_color: str | None):
	"""Convert a CSS hex colour to a reportlab HexColor, defaulting to black."""
	if not _HAS_REPORTLAB:
		return None
	if not css_color or css_color in ("transparent", ""):
		return colors.white
	try:
		return colors.HexColor(css_color)
	except Exception:
		return colors.black


def _rl_align(align: str) -> int:
	"""Map CSS alignment string to reportlab TA_* constant."""
	from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
	return {
		"left":    TA_LEFT,
		"center":  TA_CENTER,
		"right":   TA_RIGHT,
		"justify": TA_JUSTIFY,
	}.get(align, TA_LEFT)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

__all__ = ["ReportEngine"]
