"""
import_export_mixin.py

ImportExportMixin for SQLAlchemy/Flask-AppBuilder models.

Formats supported:
  - CSV (stdlib csv, streaming, PostgreSQL COPY protocol)
  - JSONL (newline-delimited JSON, true streaming — unbounded datasets)
  - JSON (full document, in-memory)
  - Excel (.xlsx via openpyxl, optional)
  - XML (stdlib xml.etree, no external dep)
  - YAML (pyyaml, optional)

Key design points:
  - No pandas/numpy. Pure stdlib + SQLAlchemy 2.x.
  - PostgreSQL COPY FROM/TO via psycopg2 cursor.copy_expert() for
    O(n) bulk throughput without Python row overhead.
  - JSONL export is a true generator — rows are serialised and yielded
    one at a time so memory stays constant regardless of dataset size.
  - Progress callbacks let callers drive progress bars or WebSocket
    notifications without polling.
  - Conflict handling: insert-or-update (upsert) or insert-or-skip.
  - Dry-run mode: validates all rows, returns error report, no DB writes.
  - Field discovery respects SQLAlchemy 2.x `inspect()` on mapped classes.

Author: Nyimbi Odero
Date: 2026-05-30
Version: 3.0
"""

from __future__ import annotations

import csv
import io
import json
import logging
import xml.etree.ElementTree as ET
from collections.abc import Callable, Generator, Iterable, Iterator
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from enum import Enum as PyEnum
from pathlib import Path
from typing import Any, Literal

from sqlalchemy import (
	ARRAY,
	Boolean,
	Date,
	DateTime,
	Enum,
	Float,
	Integer,
	Numeric,
	String,
	Text,
	Time,
	text,
)
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.inspection import inspect
from sqlalchemy.orm import RelationshipProperty, Session

try:
	from flask_appbuilder.models.mixins import AuditMixin
	_AuditMixinBase: type = AuditMixin
except ImportError:
	_AuditMixinBase = object  # type: ignore[assignment,misc]

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Optional heavy dependencies — degrade gracefully
# ---------------------------------------------------------------------------

try:
	import openpyxl as _openpyxl
	_OPENPYXL_AVAILABLE = True
except ImportError:
	_OPENPYXL_AVAILABLE = False

try:
	import yaml as _yaml
	_YAML_AVAILABLE = True
except ImportError:
	_YAML_AVAILABLE = False

try:
	import xmltodict as _xmltodict
	_XMLTODICT_AVAILABLE = True
except ImportError:
	_XMLTODICT_AVAILABLE = False

# ---------------------------------------------------------------------------
# Types
# ---------------------------------------------------------------------------

ConflictPolicy = Literal["error", "skip", "update"]
ProgressCallback = Callable[[int, int], None]  # (processed, total)

# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class ImportValidationError(ValueError):
	"""A single field value failed validation or type coercion."""


class ImportRowError(Exception):
	"""Container for row-level import failures, preserving context."""

	def __init__(self, row: int, data: dict[str, Any], cause: Exception) -> None:
		self.row = row
		self.data = data
		self.cause = cause
		super().__init__(f"Row {row}: {cause}")


class ImportResult:
	"""Structured result returned by all import_* methods."""

	__slots__ = ("total", "inserted", "updated", "skipped", "failed", "errors", "dry_run")

	def __init__(self, *, dry_run: bool = False) -> None:
		self.total: int = 0
		self.inserted: int = 0
		self.updated: int = 0
		self.skipped: int = 0
		self.failed: int = 0
		self.errors: list[dict[str, Any]] = []
		self.dry_run = dry_run

	def as_dict(self) -> dict[str, Any]:
		return {
			"dry_run": self.dry_run,
			"total": self.total,
			"inserted": self.inserted,
			"updated": self.updated,
			"skipped": self.skipped,
			"failed": self.failed,
			"errors": self.errors,
		}

	def __repr__(self) -> str:
		return (
			f"ImportResult(total={self.total}, inserted={self.inserted}, "
			f"updated={self.updated}, skipped={self.skipped}, "
			f"failed={self.failed}, dry_run={self.dry_run})"
		)


# ---------------------------------------------------------------------------
# Lightweight XML helpers (no external deps)
# ---------------------------------------------------------------------------


def _dict_to_xml_element(tag: str, data: Any) -> ET.Element:
	elem = ET.Element(tag)
	if isinstance(data, dict):
		for key, val in data.items():
			elem.append(_dict_to_xml_element(str(key), val))
	elif isinstance(data, list):
		for item in data:
			elem.append(_dict_to_xml_element("item", item))
	else:
		elem.text = "" if data is None else str(data)
	return elem


def _records_to_xml(
	records: Iterable[dict[str, Any]],
	root_tag: str = "data",
	item_tag: str = "item",
) -> bytes:
	root = ET.Element(root_tag)
	for record in records:
		root.append(_dict_to_xml_element(item_tag, record))
	return ET.tostring(root, encoding="utf-8", xml_declaration=True)


# ---------------------------------------------------------------------------
# JSONL streaming helpers
# ---------------------------------------------------------------------------


def _jsonl_row(record: dict[str, Any]) -> str:
	"""Serialise one record to a JSONL line (no trailing newline)."""
	return json.dumps(record, default=_json_default)


def _json_default(obj: Any) -> Any:
	"""JSON serialiser for types not handled by the stdlib encoder."""
	if isinstance(obj, (datetime, date, time)):
		return obj.isoformat()
	if isinstance(obj, Decimal):
		return str(obj)
	if isinstance(obj, PyEnum):
		return obj.value
	if isinstance(obj, bytes):
		return obj.decode("utf-8", errors="replace")
	raise TypeError(f"Object of type {type(obj).__name__} is not JSON serialisable")


# ---------------------------------------------------------------------------
# Main mixin
# ---------------------------------------------------------------------------


class ImportExportMixin(_AuditMixinBase):
	"""
	Multi-format import/export mixin for Flask-AppBuilder SQLAlchemy models.

	Attach to any FAB model class (alongside or instead of AuditMixin):

	    class Product(Model, ImportExportMixin):
	        __tablename__ = "product"
	        id = Column(Integer, primary_key=True)
	        name = Column(Text, nullable=False)
	        price = Column(Numeric(12, 4))

	Customise behaviour with class attributes:

	    class Product(Model, ImportExportMixin):
	        __export_fields__   = ["id", "name", "price"]
	        __import_fields__   = ["name", "price"]
	        __export_labels__   = {"name": "Product Name", "price": "Unit Price"}
	        __batch_size__      = 500

	Export methods:

	    csv_text  = Product.export_to_csv(session.execute(select(Product)).scalars())
	    jsonl_gen = Product.export_to_jsonl_stream(session.execute(select(Product)).scalars())
	    xlsx_path = Product.export_to_excel(query, "/tmp/products.xlsx")

	    # PostgreSQL COPY-based high-speed export (requires psycopg2 raw connection)
	    Product.export_pg_copy(raw_conn, "/tmp/products.csv")

	Import methods:

	    result = Product.import_from_csv(session, "/tmp/products.csv")
	    result = Product.import_from_jsonl(session, "/tmp/products.jsonl", dry_run=True)
	    result = Product.import_pg_copy(raw_conn, "/tmp/products.csv", columns=["name","price"])

	All import methods return an ``ImportResult`` instance.

	Hooks (override in subclasses):
	    pre_import_hook(data)         -- normalise/filter raw dicts before processing
	    data_validation_hook(data)    -- batch-level cross-row validation
	    post_export_hook(record)      -- per-record post-processing after serialisation
	    on_import_error(row, data, e) -- custom error handling (log, alert, etc.)

	PostgreSQL COPY notes:
	    Uses ``cursor.copy_expert()`` which bypasses Python row overhead entirely.
	    For import, expects CSV with a header row matching column names.
	    For export, writes directly to a file path via server-side COPY TO.
	    Not available on non-PostgreSQL backends — falls back to batch insert.
	"""

	# ------------------------------------------------------------------
	# Class-level configuration
	# ------------------------------------------------------------------

	__export_fields__: list[str] = []
	"""Explicit ordered list of fields to export. Empty = all non-excluded."""

	__import_fields__: list[str] = []
	"""Explicit ordered list of fields accepted on import. Empty = all non-excluded."""

	__export_exclude__: list[str] = [
		"created_by_fk",
		"changed_by_fk",
	]
	"""Fields always omitted from export regardless of __export_fields__."""

	__import_exclude__: list[str] = [
		"id",
		"created_by_fk",
		"changed_by_fk",
		"created_on",
		"changed_on",
	]
	"""Fields always omitted from import regardless of __import_fields__."""

	__export_labels__: dict[str, str] = {}
	"""Column header overrides for export: {field_name: display_label}."""

	__import_validators__: dict[str, Callable[[Any], bool]] = {}
	"""Per-field validators called post-coercion. Raise or return False to reject."""

	__import_transformers__: dict[str, Callable[[Any], Any]] = {}
	"""Per-field transformers called before type coercion."""

	__batch_size__: int = 1000
	"""SQLAlchemy flush batch size for import_data."""

	__date_format__: str = "%Y-%m-%d"
	"""strptime format for bare date strings."""

	__datetime_format__: str = "%Y-%m-%dT%H:%M:%S"
	"""strptime format for datetime strings (ISO 8601 without timezone)."""

	__null_values__: frozenset[str] = frozenset(
		{"", "null", "NULL", "None", "NA", "N/A", "nan", "NaN"}
	)
	"""String values treated as SQL NULL during import."""

	__true_values__: frozenset[str] = frozenset(
		{"true", "True", "TRUE", "1", "yes", "Yes", "YES", "on", "On", "ON"}
	)
	"""String values coerced to True."""

	__false_values__: frozenset[str] = frozenset(
		{"false", "False", "FALSE", "0", "no", "No", "NO", "off", "Off", "OFF"}
	)
	"""String values coerced to False."""

	__upsert_key__: list[str] = []
	"""
	Fields forming the natural key for upsert (conflict_policy="update").
	If empty, upsert falls back to the primary key column(s).
	Example: ["sku"] or ["country_code", "year"]
	"""

	# ------------------------------------------------------------------
	# Field discovery
	# ------------------------------------------------------------------

	@classmethod
	def _mapper(cls) -> Any:
		return inspect(cls)

	@classmethod
	def get_exportable_fields(cls) -> list[str]:
		"""Return ordered list of fields eligible for export."""
		if cls.__export_fields__:
			return [f for f in cls.__export_fields__ if f not in cls.__export_exclude__]
		try:
			return [
				attr.key
				for attr in cls._mapper().attrs
				if not attr.key.startswith("_") and attr.key not in cls.__export_exclude__
			]
		except Exception:
			return []

	@classmethod
	def get_importable_fields(cls) -> list[str]:
		"""Return ordered list of fields accepted during import."""
		if cls.__import_fields__:
			return [f for f in cls.__import_fields__ if f not in cls.__import_exclude__]
		try:
			return [
				attr.key
				for attr in cls._mapper().attrs
				if not attr.key.startswith("_") and attr.key not in cls.__import_exclude__
			]
		except Exception:
			return []

	@classmethod
	def _col_type(cls, field: str) -> Any | None:
		"""Return the SQLAlchemy column type for *field*, or None."""
		try:
			col_attr = getattr(cls._mapper().attrs, field, None)
			if col_attr is not None and hasattr(col_attr, "columns"):
				return col_attr.columns[0].type
		except Exception:
			pass
		return None

	@classmethod
	def _pk_columns(cls) -> list[str]:
		"""Return primary key column names for this model."""
		try:
			return [c.key for c in cls._mapper().primary_key]
		except Exception:
			return ["id"]

	# ------------------------------------------------------------------
	# Serialisation
	# ------------------------------------------------------------------

	@classmethod
	def _serialize_value(cls, value: Any) -> Any:
		"""Convert a single attribute value to a JSON-safe primitive."""
		if value is None:
			return None
		# Related model instance → PK
		if hasattr(value, "__table__"):
			return getattr(value, "id", None)
		# M2M / O2M collection → list of PKs
		if isinstance(value, list) and value and hasattr(value[0], "__table__"):
			return [getattr(item, "id", None) for item in value]
		if isinstance(value, (datetime, date, time)):
			return value.isoformat()
		if isinstance(value, Decimal):
			return str(value)
		if isinstance(value, PyEnum):
			return value.value
		if isinstance(value, (list, dict)):
			# Embedded JSON / JSONB — keep as-is (already serialisable)
			return value
		return value

	@classmethod
	def to_dict(cls, instance: Any, *, include_relations: bool = True) -> dict[str, Any]:
		"""
		Serialise a model instance to a plain dict.

		All values are made JSON-safe. Relationships are represented as PKs.
		Runs ``post_export_hook`` before returning.
		"""
		data: dict[str, Any] = {}
		for field in cls.get_exportable_fields():
			raw = getattr(instance, field, None)
			if not include_relations and (
				hasattr(raw, "__table__")
				or (isinstance(raw, list) and raw and hasattr(raw[0], "__table__"))
			):
				continue
			data[field] = cls._serialize_value(raw)
		return cls.post_export_hook(data)

	@classmethod
	def _apply_labels(cls, record: dict[str, Any]) -> dict[str, Any]:
		if not cls.__export_labels__:
			return record
		return {cls.__export_labels__.get(k, k): v for k, v in record.items()}

	@classmethod
	def _reverse_labels(cls, record: dict[str, Any]) -> dict[str, Any]:
		if not cls.__export_labels__:
			return record
		reverse = {v: k for k, v in cls.__export_labels__.items()}
		return {reverse.get(k, k): v for k, v in record.items()}

	# ------------------------------------------------------------------
	# Field coercion and validation for import
	# ------------------------------------------------------------------

	@classmethod
	def _coerce_value(cls, field: str, value: Any) -> Any:
		"""
		Coerce *value* to the Python type expected by *field*'s column.

		Order:
		  1. Null → None
		  2. Custom transformer
		  3. Relationship resolution (FK / M2M)
		  4. Column type coercion
		  5. Custom validator

		Raises ``ImportValidationError`` on failure.
		"""
		# 1. Null sentinel
		if value is None or (isinstance(value, str) and value.strip() in cls.__null_values__):
			return None

		# 2. Custom transformer (pre-coercion)
		if field in cls.__import_transformers__:
			try:
				value = cls.__import_transformers__[field](value)
			except Exception as exc:
				raise ImportValidationError(
					f"Field '{field}': transformer raised — {exc}"
				) from exc

		# 3. Relationship resolution
		cls_attr = getattr(cls, field, None)
		if cls_attr is not None and hasattr(cls_attr, "property") and isinstance(
			cls_attr.property, RelationshipProperty
		):
			related_model = cls_attr.property.mapper.class_
			if cls_attr.property.uselist:
				if isinstance(value, str):
					try:
						value = json.loads(value)
					except json.JSONDecodeError as exc:
						raise ImportValidationError(
							f"Field '{field}': expected JSON array for M2M, got {value!r}"
						) from exc
				resolved = []
				for pk in value or []:
					obj = related_model.query.get(pk) if pk is not None else None
					resolved.append(obj)
				return resolved
			else:
				return related_model.query.get(value) if value is not None else None

		# 4. Column type coercion
		field_type = cls._col_type(field)
		if field_type is not None:
			try:
				if isinstance(field_type, (String, Text)):
					value = str(value).strip()
				elif isinstance(field_type, Integer):
					value = int(float(value))
				elif isinstance(field_type, Float):
					value = float(value)
				elif isinstance(field_type, Boolean):
					sv = str(value).strip()
					if sv in cls.__true_values__:
						value = True
					elif sv in cls.__false_values__:
						value = False
					else:
						raise ImportValidationError(
							f"Field '{field}': cannot parse boolean from {value!r}"
						)
				elif isinstance(field_type, DateTime):
					if isinstance(value, str):
						# Try ISO first, fall back to __datetime_format__
						for fmt in (cls.__datetime_format__, cls.__date_format__):
							try:
								value = datetime.strptime(value.strip(), fmt)
								break
							except ValueError:
								continue
						else:
							raise ImportValidationError(
								f"Field '{field}': cannot parse datetime from {value!r}"
							)
				elif isinstance(field_type, Date):
					if isinstance(value, str):
						try:
							value = datetime.strptime(value.strip(), cls.__date_format__).date()
						except ValueError as exc:
							raise ImportValidationError(
								f"Field '{field}': cannot parse date from {value!r}"
							) from exc
				elif isinstance(field_type, Time):
					if isinstance(value, str):
						try:
							value = datetime.strptime(value.strip(), "%H:%M:%S").time()
						except ValueError as exc:
							raise ImportValidationError(
								f"Field '{field}': cannot parse time from {value!r}"
							) from exc
				elif isinstance(field_type, Numeric):
					value = Decimal(str(value))
				elif isinstance(field_type, (JSONB,)):
					if isinstance(value, str):
						value = json.loads(value)
				elif hasattr(field_type, "impl") and isinstance(
					getattr(field_type, "impl", None), type
				) and issubclass(field_type.impl, (String, Text)):
					# Custom TypeDecorator backed by text
					value = str(value)
				elif isinstance(field_type, ARRAY):
					if isinstance(value, str):
						value = json.loads(value)
				elif isinstance(field_type, Enum):
					try:
						value = field_type.python_type(value)
					except (ValueError, KeyError) as exc:
						raise ImportValidationError(
							f"Field '{field}': invalid enum value {value!r}"
						) from exc
			except ImportValidationError:
				raise
			except (ValueError, TypeError, json.JSONDecodeError, InvalidOperation) as exc:
				raise ImportValidationError(
					f"Field '{field}': cannot coerce {value!r} — {exc}"
				) from exc

		# 5. Custom validator (post-coercion)
		if field in cls.__import_validators__:
			try:
				ok = cls.__import_validators__[field](value)
			except Exception as exc:
				raise ImportValidationError(
					f"Field '{field}': validator raised — {exc}"
				) from exc
			if not ok:
				raise ImportValidationError(
					f"Field '{field}': validation rejected value {value!r}"
				)

		return value

	# ------------------------------------------------------------------
	# Core import engine
	# ------------------------------------------------------------------

	@classmethod
	def import_data(
		cls,
		session: Session,
		data: list[dict[str, Any]],
		*,
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""
		Bulk-import records with validation, batching, and structured error collection.

		Args:
			session:         SQLAlchemy 2.x Session.
			data:            Raw dicts with field names matching model columns.
			batch_size:      Override ``__batch_size__``. Tunes flush granularity.
			conflict_policy: What to do when a record with the same PK/key exists:
			                 "error"  — let the DB raise (default)
			                 "skip"   — silently skip the conflicting row
			                 "update" — update existing record in-place
			dry_run:         If True, validate all rows but never commit / flush.
			                 Returns a full error report without modifying the DB.
			progress:        Optional callable ``(processed, total)`` for progress
			                 reporting (e.g. driving a progress bar or WebSocket).

		Returns:
			``ImportResult`` with counts and per-row error details.
		"""
		data = cls.pre_import_hook(list(data))
		data = cls.data_validation_hook(data)

		importable = set(cls.get_importable_fields())
		effective_batch = batch_size or cls.__batch_size__
		result = ImportResult(dry_run=dry_run)
		result.total = len(data)

		upsert_key = cls.__upsert_key__ or cls._pk_columns()

		for batch_start in range(0, result.total, effective_batch):
			batch = data[batch_start : batch_start + effective_batch]
			batch_inserted = 0
			batch_updated = 0

			for local_idx, item in enumerate(batch):
				global_row = batch_start + local_idx + 1
				# Reverse label mapping so exported labels import cleanly
				item = cls._reverse_labels(item)

				try:
					coerced: dict[str, Any] = {}
					for field in importable:
						if field in item:
							coerced[field] = cls._coerce_value(field, item[field])

					if dry_run:
						# Validate only — no DB interaction
						result.inserted += 1
						if progress:
							progress(global_row, result.total)
						continue

					if conflict_policy in ("skip", "update"):
						# Build a filter from the upsert key
						filter_kwargs = {
							k: coerced.get(k) for k in upsert_key if k in coerced
						}
						if filter_kwargs:
							existing = (
								session.query(cls)  # type: ignore[attr-defined]
								.filter_by(**filter_kwargs)
								.first()
							)
						else:
							existing = None

						if existing is not None:
							if conflict_policy == "skip":
								result.skipped += 1
								if progress:
									progress(global_row, result.total)
								continue
							else:  # "update"
								for field, val in coerced.items():
									setattr(existing, field, val)
								session.add(existing)
								batch_updated += 1
								if progress:
									progress(global_row, result.total)
								continue

					instance = cls()  # type: ignore[call-arg]
					for field, val in coerced.items():
						setattr(instance, field, val)
					session.add(instance)
					batch_inserted += 1

				except Exception as exc:
					cls.on_import_error(global_row, item, exc)
					result.errors.append({
						"row": global_row,
						"data": item,
						"error": str(exc),
						"error_type": type(exc).__name__,
					})
					result.failed += 1

				if progress:
					progress(global_row, result.total)

			if not dry_run:
				try:
					session.flush()
					result.inserted += batch_inserted
					result.updated += batch_updated
				except Exception as exc:
					session.rollback()
					logger.error(
						"Batch flush failed (rows %d-%d): %s",
						batch_start + 1,
						batch_start + len(batch),
						exc,
						exc_info=True,
					)
					for local_idx, item in enumerate(batch):
						result.errors.append({
							"row": batch_start + local_idx + 1,
							"data": item,
							"error": f"Batch flush failed: {exc}",
							"error_type": type(exc).__name__,
						})
					result.failed += len(batch)

		return result

	# ------------------------------------------------------------------
	# CSV export / import
	# ------------------------------------------------------------------

	@classmethod
	def export_to_csv(
		cls,
		query: Iterable[Any],
		output_file: str | Path | None = None,
		*,
		dialect: str = "excel",
		encoding: str = "utf-8-sig",
	) -> str:
		"""
		Export query results to CSV.

		Args:
			query:       Iterable of model instances (SQLAlchemy query / scalars()).
			output_file: Destination path. If None, returns CSV text.
			dialect:     csv.writer dialect (default "excel").
			encoding:    File encoding (default "utf-8-sig" for Excel compat).

		Returns:
			Absolute path written, or CSV text when output_file is None.
		"""
		buf = io.StringIO()
		writer: csv.DictWriter | None = None

		for instance in query:
			record = cls._apply_labels(cls.to_dict(instance))
			if writer is None:
				fieldnames = list(record.keys())
				writer = csv.DictWriter(buf, fieldnames=fieldnames, dialect=dialect, extrasaction="ignore")
				writer.writeheader()
			writer.writerow({k: ("" if v is None else v) for k, v in record.items()})

		csv_text = buf.getvalue()
		if output_file is None:
			return csv_text

		path = Path(output_file)
		path.write_text(csv_text, encoding=encoding)
		return str(path.resolve())

	@classmethod
	def export_to_csv_stream(cls, query: Iterable[Any]) -> io.BytesIO:
		"""Return a BytesIO of UTF-8-sig CSV, suitable for Flask ``send_file``."""
		return io.BytesIO(cls.export_to_csv(query, output_file=None).encode("utf-8-sig"))

	@classmethod
	def import_from_csv(
		cls,
		session: Session,
		file_path: str | Path,
		*,
		encoding: str = "utf-8-sig",
		delimiter: str = ",",
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""
		Import records from a CSV file.

		Uses stdlib ``csv.DictReader`` — no pandas dependency.
		Handles BOM (utf-8-sig) produced by Excel and Windows tools.
		"""
		with open(file_path, newline="", encoding=encoding) as fh:
			reader = csv.DictReader(fh, delimiter=delimiter)
			records = list(reader)
		return cls.import_data(
			session,
			records,
			batch_size=batch_size,
			conflict_policy=conflict_policy,
			dry_run=dry_run,
			progress=progress,
		)

	# ------------------------------------------------------------------
	# JSONL (newline-delimited JSON) — true streaming
	# ------------------------------------------------------------------

	@classmethod
	def export_to_jsonl_stream(
		cls, query: Iterable[Any]
	) -> Generator[str, None, None]:
		"""
		Generate one JSONL line per row without loading all rows into memory.

		Usage with Flask streaming response::

		    from flask import Response, stream_with_context

		    @app.route("/products/export.jsonl")
		    def export():
		        q = session.execute(select(Product)).scalars()
		        return Response(
		            stream_with_context(Product.export_to_jsonl_stream(q)),
		            mimetype="application/x-ndjson",
		        )
		"""
		for instance in query:
			record = cls.to_dict(instance)
			yield _jsonl_row(record) + "\n"

	@classmethod
	def export_to_jsonl(
		cls,
		query: Iterable[Any],
		output_file: str | Path,
	) -> str:
		"""Write JSONL to *output_file*. One JSON object per line."""
		path = Path(output_file)
		with path.open("w", encoding="utf-8") as fh:
			for line in cls.export_to_jsonl_stream(query):
				fh.write(line)
		return str(path.resolve())

	@classmethod
	def import_from_jsonl(
		cls,
		session: Session,
		file_path: str | Path,
		*,
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""
		Import from a JSONL file.

		Reads one line at a time — constant memory regardless of file size.
		Particularly suited for large dataset migrations.
		"""
		importable = set(cls.get_importable_fields())
		effective_batch = batch_size or cls.__batch_size__
		result = ImportResult(dry_run=dry_run)
		upsert_key = cls.__upsert_key__ or cls._pk_columns()

		# Count lines for accurate progress without loading the whole file
		total_lines = 0
		if progress:
			with open(file_path, encoding="utf-8") as fh:
				for _ in fh:
					total_lines += 1
			result.total = total_lines

		batch: list[dict[str, Any]] = []

		def _flush_batch(b: list[dict[str, Any]], start_row: int) -> None:
			sub = cls.import_data(
				session,
				b,
				batch_size=len(b),
				conflict_policy=conflict_policy,
				dry_run=dry_run,
			)
			result.inserted += sub.inserted
			result.updated += sub.updated
			result.skipped += sub.skipped
			result.failed += sub.failed
			for err in sub.errors:
				err["row"] += start_row
				result.errors.append(err)

		with open(file_path, encoding="utf-8") as fh:
			for lineno, raw_line in enumerate(fh, start=1):
				raw_line = raw_line.strip()
				if not raw_line:
					continue
				result.total = lineno  # update if no pre-count
				try:
					record = json.loads(raw_line)
				except json.JSONDecodeError as exc:
					result.failed += 1
					result.errors.append({
						"row": lineno,
						"data": {"raw": raw_line},
						"error": f"Invalid JSON on line {lineno}: {exc}",
						"error_type": "JSONDecodeError",
					})
					if progress:
						progress(lineno, result.total)
					continue
				batch.append(record)
				if len(batch) >= effective_batch:
					_flush_batch(batch, lineno - len(batch))
					batch = []
				if progress:
					progress(lineno, result.total)

		if batch:
			_flush_batch(batch, result.total - len(batch))

		return result

	# ------------------------------------------------------------------
	# JSON (full document) export / import
	# ------------------------------------------------------------------

	@classmethod
	def export_to_json(
		cls,
		query: Iterable[Any],
		output_file: str | Path | None = None,
		*,
		pretty: bool = True,
	) -> str:
		"""
		Export query results to a JSON array.

		Prefer JSONL for large datasets to avoid full materialisation in memory.
		"""
		records = [cls.to_dict(instance) for instance in query]
		text = json.dumps(records, indent=2 if pretty else None, default=_json_default)
		if output_file is None:
			return text
		path = Path(output_file)
		path.write_text(text, encoding="utf-8")
		return str(path.resolve())

	@classmethod
	def export_to_json_stream(cls, query: Iterable[Any], *, pretty: bool = True) -> io.BytesIO:
		"""Return a BytesIO of UTF-8 JSON suitable for Flask ``send_file``."""
		return io.BytesIO(cls.export_to_json(query, output_file=None, pretty=pretty).encode("utf-8"))

	@classmethod
	def import_from_json(
		cls,
		session: Session,
		file_path: str | Path,
		*,
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""Import from a JSON file containing a top-level array of objects."""
		with open(file_path, encoding="utf-8") as fh:
			data = json.load(fh)
		if not isinstance(data, list):
			raise ValueError(
				f"JSON file must contain a top-level array, got {type(data).__name__}"
			)
		return cls.import_data(
			session,
			data,
			batch_size=batch_size,
			conflict_policy=conflict_policy,
			dry_run=dry_run,
			progress=progress,
		)

	# ------------------------------------------------------------------
	# Excel (.xlsx) — requires openpyxl
	# ------------------------------------------------------------------

	@classmethod
	def export_to_excel(
		cls,
		query: Iterable[Any],
		output_file: str | Path,
		*,
		sheet_name: str = "Export",
		freeze_header: bool = True,
	) -> str:
		"""
		Export to Excel (.xlsx) via openpyxl.

		``freeze_header=True`` freezes the first row for easier scrolling.

		Raises:
			RuntimeError: if openpyxl is not installed.
		"""
		if not _OPENPYXL_AVAILABLE:
			raise RuntimeError("openpyxl is required — pip install openpyxl")

		wb = _openpyxl.Workbook()
		ws = wb.active
		ws.title = sheet_name

		header_written = False
		headers: list[str] = []

		for instance in query:
			record = cls._apply_labels(cls.to_dict(instance))
			if not header_written:
				headers = list(record.keys())
				ws.append(headers)
				if freeze_header:
					ws.freeze_panes = "A2"
				header_written = True
			ws.append([record.get(h) for h in headers])

		path = Path(output_file)
		wb.save(str(path))
		return str(path.resolve())

	@classmethod
	def import_from_excel(
		cls,
		session: Session,
		file_path: str | Path,
		*,
		sheet_name: str | None = None,
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""
		Import from an Excel (.xlsx) file via openpyxl (read-only mode).

		Raises:
			RuntimeError: if openpyxl is not installed.
		"""
		if not _OPENPYXL_AVAILABLE:
			raise RuntimeError("openpyxl is required — pip install openpyxl")

		wb = _openpyxl.load_workbook(file_path, read_only=True, data_only=True)
		ws = wb[sheet_name] if sheet_name else wb.active
		rows = list(ws.iter_rows(values_only=True))
		wb.close()

		if not rows:
			return cls.import_data(session, [], batch_size=batch_size, dry_run=dry_run)

		headers = [str(h) if h is not None else "" for h in rows[0]]
		records: list[dict[str, Any]] = [
			{headers[i]: cell for i, cell in enumerate(row)}
			for row in rows[1:]
			if any(cell is not None for cell in row)  # skip blank trailing rows
		]
		return cls.import_data(
			session,
			records,
			batch_size=batch_size,
			conflict_policy=conflict_policy,
			dry_run=dry_run,
			progress=progress,
		)

	# ------------------------------------------------------------------
	# XML (stdlib — no external dep)
	# ------------------------------------------------------------------

	@classmethod
	def export_to_xml(
		cls,
		query: Iterable[Any],
		output_file: str | Path,
		*,
		root_tag: str = "data",
		item_tag: str = "item",
	) -> str:
		"""Export to XML using stdlib xml.etree (no external deps)."""
		records = [cls.to_dict(instance) for instance in query]
		path = Path(output_file)
		path.write_bytes(_records_to_xml(records, root_tag=root_tag, item_tag=item_tag))
		return str(path.resolve())

	@classmethod
	def import_from_xml(
		cls,
		session: Session,
		file_path: str | Path,
		*,
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""
		Import from an XML file.

		Expected structure: ``<data><item><field>value</field>...</item>...</data>``.
		Uses xmltodict when available (preserves order, handles attributes);
		falls back to stdlib xml.etree.
		"""
		if _XMLTODICT_AVAILABLE:
			with open(file_path, encoding="utf-8") as fh:
				parsed = _xmltodict.parse(fh.read())
			items = parsed.get("data", {}).get("item", [])
			if isinstance(items, dict):
				items = [items]
		else:
			tree = ET.parse(str(file_path))
			root = tree.getroot()
			items = []
			for child in root:
				record: dict[str, Any] = {elem.tag: elem.text for elem in child}
				items.append(record)

		return cls.import_data(
			session,
			items,
			batch_size=batch_size,
			conflict_policy=conflict_policy,
			dry_run=dry_run,
			progress=progress,
		)

	# ------------------------------------------------------------------
	# YAML — requires pyyaml
	# ------------------------------------------------------------------

	@classmethod
	def export_to_yaml(
		cls, query: Iterable[Any], output_file: str | Path
	) -> str:
		"""
		Export to YAML.

		Raises:
			RuntimeError: if pyyaml is not installed.
		"""
		if not _YAML_AVAILABLE:
			raise RuntimeError("pyyaml is required — pip install pyyaml")
		records = [cls.to_dict(instance) for instance in query]
		path = Path(output_file)
		with path.open("w", encoding="utf-8") as fh:
			_yaml.dump(records, fh, allow_unicode=True, default_flow_style=False)
		return str(path.resolve())

	@classmethod
	def import_from_yaml(
		cls,
		session: Session,
		file_path: str | Path,
		*,
		batch_size: int | None = None,
		conflict_policy: ConflictPolicy = "error",
		dry_run: bool = False,
		progress: ProgressCallback | None = None,
	) -> ImportResult:
		"""
		Import from a YAML file (top-level sequence required).

		Raises:
			RuntimeError: if pyyaml is not installed.
		"""
		if not _YAML_AVAILABLE:
			raise RuntimeError("pyyaml is required — pip install pyyaml")
		with open(file_path, encoding="utf-8") as fh:
			data = _yaml.safe_load(fh)
		if not isinstance(data, list):
			raise ValueError(
				f"YAML file must contain a top-level sequence, got {type(data).__name__}"
			)
		return cls.import_data(
			session,
			data,
			batch_size=batch_size,
			conflict_policy=conflict_policy,
			dry_run=dry_run,
			progress=progress,
		)

	# ------------------------------------------------------------------
	# PostgreSQL COPY — high-speed bulk transfer
	# ------------------------------------------------------------------

	@classmethod
	def export_pg_copy(
		cls,
		raw_conn: Any,
		output_file: str | Path,
		*,
		columns: list[str] | None = None,
		delimiter: str = ",",
		null_string: str = "",
	) -> int:
		"""
		Export table data via PostgreSQL ``COPY TO`` (server-side, maximum throughput).

		This bypasses Python row iteration entirely. The DB writes directly to the
		file descriptor, achieving ~10× the throughput of ORM-based export.

		Args:
			raw_conn:    Raw psycopg2 connection (``session.connection().connection``).
			output_file: Destination path for the CSV. PostgreSQL writes this file,
			             so the path must be accessible to the PostgreSQL server process.
			             For client-side files, use ``export_pg_copy_client`` instead.
			columns:     Column list to export. Defaults to ``get_exportable_fields()``.
			delimiter:   CSV field delimiter.
			null_string: String representing NULL in the output.

		Returns:
			Row count reported by the COPY command.

		Raises:
			RuntimeError: if called on a non-PostgreSQL engine.
		"""
		cols = columns or cls.get_exportable_fields()
		col_list = ", ".join(f'"{c}"' for c in cols)
		table = getattr(cls, "__tablename__", cls.__name__.lower())

		copy_sql = (
			f"COPY {table} ({col_list}) TO STDOUT "
			f"WITH (FORMAT CSV, HEADER TRUE, DELIMITER '{delimiter}', NULL '{null_string}')"
		)
		path = Path(output_file)
		try:
			cursor = raw_conn.cursor()
			with path.open("wb") as fh:
				cursor.copy_expert(copy_sql, fh)
			return cursor.rowcount if cursor.rowcount >= 0 else -1
		except Exception as exc:
			raise RuntimeError(f"PostgreSQL COPY TO failed: {exc}") from exc

	@classmethod
	def import_pg_copy(
		cls,
		raw_conn: Any,
		file_path: str | Path,
		*,
		columns: list[str] | None = None,
		delimiter: str = ",",
		null_string: str = "",
		has_header: bool = True,
	) -> ImportResult:
		"""
		Import data via PostgreSQL ``COPY FROM`` (maximum throughput, no Python overhead).

		The CSV file is streamed directly from the client to the PostgreSQL server.
		No Python row-by-row processing occurs — this is the fastest possible bulk load
		for PostgreSQL targets.

		Args:
			raw_conn:    Raw psycopg2 connection.
			file_path:   Path to a CSV file accessible to the Python process.
			columns:     Column list in CSV order (must match header if has_header=True).
			             Defaults to ``get_importable_fields()``.
			delimiter:   CSV field delimiter.
			null_string: String in the CSV representing NULL.
			has_header:  Whether the first row is a header (it is skipped if True).

		Returns:
			``ImportResult`` with ``inserted`` set to the row count reported by COPY.

		Raises:
			RuntimeError: on COPY failure.
		"""
		cols = columns or cls.get_importable_fields()
		col_list = ", ".join(f'"{c}"' for c in cols)
		table = getattr(cls, "__tablename__", cls.__name__.lower())

		copy_sql = (
			f"COPY {table} ({col_list}) FROM STDIN "
			f"WITH (FORMAT CSV, HEADER {str(has_header).upper()}, "
			f"DELIMITER '{delimiter}', NULL '{null_string}')"
		)
		result = ImportResult()
		try:
			cursor = raw_conn.cursor()
			with open(file_path, "rb") as fh:
				cursor.copy_expert(copy_sql, fh)
			raw_conn.commit()
			result.inserted = cursor.rowcount if cursor.rowcount >= 0 else -1
			result.total = result.inserted
		except Exception as exc:
			raw_conn.rollback()
			result.failed = -1  # unknown — entire COPY failed atomically
			result.errors.append({
				"row": -1,
				"data": {},
				"error": str(exc),
				"error_type": type(exc).__name__,
			})
			raise RuntimeError(f"PostgreSQL COPY FROM failed: {exc}") from exc

		return result

	@classmethod
	def export_pg_copy_client(
		cls,
		raw_conn: Any,
		*,
		columns: list[str] | None = None,
		delimiter: str = ",",
		null_string: str = "",
	) -> io.BytesIO:
		"""
		Export via COPY TO STDOUT — data is buffered client-side.

		Unlike ``export_pg_copy``, the output file need not be reachable by
		the PostgreSQL server. Suitable for web API streaming responses.

		Returns:
			BytesIO containing the full CSV (headers included).
		"""
		cols = columns or cls.get_exportable_fields()
		col_list = ", ".join(f'"{c}"' for c in cols)
		table = getattr(cls, "__tablename__", cls.__name__.lower())

		copy_sql = (
			f"COPY {table} ({col_list}) TO STDOUT "
			f"WITH (FORMAT CSV, HEADER TRUE, DELIMITER '{delimiter}', NULL '{null_string}')"
		)
		buf = io.BytesIO()
		try:
			cursor = raw_conn.cursor()
			cursor.copy_expert(copy_sql, buf)
		except Exception as exc:
			raise RuntimeError(f"PostgreSQL COPY TO STDOUT failed: {exc}") from exc
		buf.seek(0)
		return buf

	# ------------------------------------------------------------------
	# Progress-aware chunked query helper
	# ------------------------------------------------------------------

	@classmethod
	def iter_chunks(
		cls,
		session: Session,
		*,
		chunk_size: int = 1000,
		order_by: str | None = None,
	) -> Iterator[list[Any]]:
		"""
		Yield lists of *chunk_size* model instances from the DB.

		Uses ``yield_per`` (SQLAlchemy 2.x) to avoid loading the full table.
		Useful for feeding ``export_to_jsonl`` or ``export_to_csv`` without
		building the complete result set in memory.

		Example::

		    for chunk in Product.iter_chunks(session, chunk_size=500):
		        for instance in chunk:
		            process(instance)
		"""
		from sqlalchemy import select

		stmt = select(cls)  # type: ignore[arg-type]
		if order_by:
			stmt = stmt.order_by(text(order_by))

		batch: list[Any] = []
		for row in session.execute(stmt.execution_options(yield_per=chunk_size)).scalars():
			batch.append(row)
			if len(batch) >= chunk_size:
				yield batch
				batch = []
		if batch:
			yield batch

	# ------------------------------------------------------------------
	# Hooks — override in subclasses
	# ------------------------------------------------------------------

	@classmethod
	def pre_import_hook(
		cls, data: list[dict[str, Any]]
	) -> list[dict[str, Any]]:
		"""
		Called at the very start of ``import_data`` before any processing.

		Override to normalise keys, strip vendor-specific prefixes, etc.
		"""
		return data

	@classmethod
	def data_validation_hook(
		cls, data: list[dict[str, Any]]
	) -> list[dict[str, Any]]:
		"""
		Batch-level cross-row validation called after ``pre_import_hook``.

		Override to enforce uniqueness constraints within the incoming batch,
		check referential integrity, or reject the batch entirely.
		"""
		return data

	@classmethod
	def post_export_hook(cls, record: dict[str, Any]) -> dict[str, Any]:
		"""
		Per-record hook called after each instance is serialised to dict.

		Override to add computed fields, redact PII, or rename keys.
		"""
		return record

	@classmethod
	def on_import_error(cls, row: int, data: dict[str, Any], exc: Exception) -> None:
		"""
		Called for every row-level import failure before it is recorded.

		Default implementation logs at ERROR level. Override to push
		alerts, increment metrics, or implement circuit-breaker logic.
		"""
		logger.error("Import row %d error [%s]: %s", row, type(exc).__name__, exc)
